import json
from typing import Any
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(data: str, filename: str = "reporte.xlsx") -> dict[str, Any]:
    """
    Genera un archivo excel.

    Args:
        data: String JSON con la estructura:
              {"sheets": [{"name": "Sheet1", "headers": [...], "rows": [[...]]}]}
        filename: Nombre del archivo Excel a generar

    Returns:
        Dict con el archivo en base64 y metadata
    """
    try:
        excel_data = json.loads(data)
        
        print("Data JSON String: ", excel_data)

        wb = Workbook()
        # remover hoja por defecto
        wb.remove(wb.active)  # type: ignore

        # Seleccionar tema
        color_scheme = {
            "header_fill": "366092",  # azul oscuro
            "header_font": "FFFFFF",  # blanco
            "alt_row_fill": "D9E2F3",  # azul claro
            "border_color": "B4C7E7",  # azul medio
        }

        # cada hoja
        for sheet_data in excel_data.get("sheets", []):
            sheet_name = sheet_data.get("name", "Sheet1")
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            # crear hoja
            ws = wb.create_sheet(title=sheet_name)

            # estilos para bordes
            thin_border = Border(
                left=Side(style="thin", color=color_scheme["border_color"]),
                right=Side(style="thin", color=color_scheme["border_color"]),
                top=Side(style="thin", color=color_scheme["border_color"]),
                bottom=Side(style="thin", color=color_scheme["border_color"]),
            )

            # escribir encabezados con formato
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12, color=color_scheme["header_font"])
                cell.fill = PatternFill(
                    start_color=color_scheme["header_fill"],
                    end_color=color_scheme["header_fill"],
                    fill_type="solid",
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = thin_border

            # escribir datos con formato alternado
            for row_num, row_data in enumerate(rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value

                    # formato para filas alternas
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(
                            start_color=color_scheme["alt_row_fill"],
                            end_color=color_scheme["alt_row_fill"],
                            fill_type="solid",
                        )

                    # alineacion segun tipo de dato
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        # formato de numero con separadores de miles
                        if isinstance(value, float):
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"
                    else:
                        cell.alignment = Alignment(horizontal="left")

                    cell.border = thin_border
                    cell.font = Font(size=11)

            # ajustar ancho de columnas automáticamente
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0

                # calcular ancho maximo
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass

                # ancho (minimo 12, maximo 50)
                adjusted_width = min(max(max_length + 2, 12), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # fijar primera fila (encabezados)
            ws.freeze_panes = "A2"

            # agregar filtros automáticos
            if headers:
                ws.auto_filter.ref = ws.dimensions

            output = BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            excel_base64 = base64.b64encode(excel_bytes).decode("utf-8")

        result: dict[str, Any] = {
            "type": "excel_file",
            "filename": filename,
            "content": excel_base64,  # type: ignore
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size_bytes": len(excel_bytes),  # type: ignore
            "success": True,
            "sheets_count": len(excel_data.get("sheets", [])),
            "message": f"Excel '{filename}' generado exitosamente con {len(excel_data.get('sheets', []))} hoja(s)",
        }
        print("Excel generado base64: ", result)

        return result
    except json.JSONDecodeError as e:
        return {
            "type": "error",
            "success": False,
            "message": f"Error parseando datos JSON: {str(e)}",
        }
    except Exception as e:
        return {
            "type": "error",
            "success": False,
            "message": f"Error generando Excel: {str(e)}",
        }
